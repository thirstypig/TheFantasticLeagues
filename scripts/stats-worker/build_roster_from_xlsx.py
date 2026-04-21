# build_roster_from_xlsx.py
"""
Builds OGBA roster_ogba_2025.csv automatically from the Excel workbook.

Assumptions (matching your current file):
- Workbook: Fantasy_Baseball_2025 - FINAL.xlsx
- Sheet: "Draft"
- Rows 2-3: team names across columns
- Rows 6-19: hitters (1B, 2B, SS, 3B, OF, C, CM, MI, DH)
- Rows 21-25: pitchers (P)
- First 4 teams use columns 2,4,6,8 (names), second 4 use 11,13,15,17.

Output:
- roster_ogba_2025.csv in the same folder.
"""

import csv
import os

from openpyxl import load_workbook

# ---- CONFIG ----

EXCEL_PATH = "Fantasy_Baseball_2025 - FINAL.xlsx"
SHEET_NAME = "Draft"
OUTPUT_CSV = "roster_ogba_2025.csv"

# Map full team names -> OGBA 3-letter codes
TEAM_CODE_MAP = {
    "Diamond Kings": "DMK",
    "Demolition Lumber Co.": "DLC",
    "Dodger Dawgs": "DDG",
    "Skunk Dogs": "SKD",
    "RGing Sluggers": "RGS",   # adjust to your exact sheet name if needed
    "Devil Dawgs": "DVD",
    "Los Doyers": "LDY",
    "The Show": "TSH",
}


def build_roster():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(
            f"Could not find {EXCEL_PATH}. Put it in this folder or update EXCEL_PATH."
        )

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet {SHEET_NAME!r} not found in workbook.")

    ws = wb[SHEET_NAME]

    # 1) Detect team columns + names from rows 2 and 3
    team_cols = []
    for col in range(2, 19):
        r2 = ws.cell(row=2, column=col).value
        r3 = ws.cell(row=3, column=col).value
        if r2 or r3:
            if r2 and r3:
                name = f"{r2} {r3}"
            elif r2:
                name = r2
            else:
                name = r3
            team_cols.append({"index": len(team_cols), "col": col, "name": str(name).strip()})

    if len(team_cols) != 8:
        print("WARNING: Expected 8 teams, found", len(team_cols))
    else:
        print("Detected teams:")
        for t in team_cols:
            print(f"  {t['index']}: col {t['col']} -> {t['name']}")

    rows_out = []

    # 2) Iterate roster rows
    # Hitters: 6–19, Pitchers: 21–25 (we’ll just scan 6–30 and use the position columns)
    for row in range(6, 30):
        pos_block1 = ws.cell(row=row, column=1).value   # for first 4 teams
        pos_block2 = ws.cell(row=row, column=10).value  # for last 4 teams

        # Skip completely empty lines
        if not pos_block1 and not pos_block2:
            continue

        for t in team_cols:
            idx = t["index"]
            team_name = t["name"]
            col = t["col"]

            player_name = ws.cell(row=row, column=col).value

            # Skip empty cells and cells that are actually just position markers
            if not player_name:
                continue
            if str(player_name).strip() in ("P", "OF", "1B", "2B", "SS", "3B", "C", "CM", "MI", "DH"):
                continue

            # Decide which position block to use (first 4 vs second 4 teams)
            if idx <= 3:
                pos = pos_block1
            else:
                pos = pos_block2

            if not pos:
                # If for some reason there's no position here, just skip
                continue

            pos_str = str(pos).strip()
            role = "P" if pos_str == "P" else "H"

            rows_out.append(
                {
                    "ogba_team_code": TEAM_CODE_MAP.get(team_name, ""),
                    "ogba_team_name": team_name,
                    "status": "act",  # default everything to 'active' for now
                    "role": role,
                    "player_name": str(player_name).strip(),
                    "mlb_id": "",     # to be filled later
                    "positions": pos_str,
                }
            )

    # 3) Write CSV
    fieldnames = [
        "ogba_team_code",
        "ogba_team_name",
        "status",
        "role",
        "player_name",
        "mlb_id",
        "positions",
    ]

    with open(OUTPUT_CSV, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows_out:
            writer.writerow(row)

    print(f"Wrote {len(rows_out)} rows to {OUTPUT_CSV}")


if __name__ == "__main__":
    build_roster()
